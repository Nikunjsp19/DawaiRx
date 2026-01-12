"""Excel report generation"""

from pathlib import Path
from typing import Dict, Any, Optional
import pandas as pd
import logging

logger = logging.getLogger(__name__)


def create_audit_report(
    output_path: str,
    reconciled_df: pd.DataFrame,
    issues_df: pd.DataFrame,
    summary: Dict[str, Any],
    dawairx_report: Optional[pd.DataFrame] = None
):
    """
    Create Excel workbook with DawaiRx format matching UI display.
    
    Args:
        output_path: Path to output Excel file
        reconciled_df: Reconciled inventory DataFrame (legacy, kept for compatibility)
        issues_df: Issues DataFrame
        summary: Summary statistics dictionary
        dawairx_report: DawaiRx format report DataFrame (matches UI display)
    """
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # If DawaiRx report is provided, use it (matches UI format)
        if dawairx_report is not None and len(dawairx_report) > 0:
            # Main sheet: Inventory Report (DawaiRx format - matches UI)
            dawairx_export = dawairx_report.copy()
            
            # Replace NaN/NA with empty string for Excel (matches UI display where 0 values are blank)
            dawairx_export = dawairx_export.fillna('')
            # Replace 0 values with empty string to match UI (DawaiRx shows blank for 0)
            for col in dawairx_export.columns:
                if dawairx_export[col].dtype in ['float64', 'int64', 'Int64', 'Float64']:
                    dawairx_export[col] = dawairx_export[col].replace(0, '')
                    dawairx_export[col] = dawairx_export[col].replace(0.0, '')
            
            dawairx_export.to_excel(writer, sheet_name="Inventory Report", index=False)
            
            # Format the main sheet
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            worksheet = writer.sheets["Inventory Report"]
            
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format numeric columns (right align)
            numeric_columns = ['RANK', 'PKG SIZE', 'TOTAL ORDERED-O', 'TOTAL BILLED-B', 
                             'TOTAL SHORTAGE-S', 'HIGHEST SHORTAGE-S', 'AMOUNT', 'COST']
            for col_idx, col_name in enumerate(dawairx_export.columns, start=1):
                if any(numeric_col in str(col_name).upper() for numeric_col in numeric_columns):
                    column_letter = get_column_letter(col_idx)
                    for row in range(2, len(dawairx_export) + 2):
                        cell = worksheet[f"{column_letter}{row}"]
                        if cell.value != '':
                            cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # Format SHORTAGE columns (red for negative values)
            shortage_cols = [col for col in dawairx_export.columns if 'SHORTAGE' in str(col).upper()]
            for col_name in shortage_cols:
                col_idx = list(dawairx_export.columns).index(col_name) + 1
                column_letter = get_column_letter(col_idx)
                for row in range(2, len(dawairx_export) + 2):
                    cell = worksheet[f"{column_letter}{row}"]
                    if cell.value != '' and isinstance(cell.value, (int, float)) and cell.value < 0:
                        cell.font = Font(color="FF0000", bold=True)
        else:
            # Fallback to legacy format if DawaiRx report not available
            logger.warning("DawaiRx report not provided, using legacy format")
            reconciled_df.to_excel(writer, sheet_name="Inventory Report", index=False)
        
        # Sheet 2: Summary
        summary_data = {
            "Metric": [
                "Total Medicines",
                "Total Ordered",
                "Total Sold",
                "Total Remaining",
                "Total Shortage",
                "Total Leftover",
                "Medicines with Shortage",
                "Medicines with Leftover",
                "Sold Percentage",
            ],
            "Value": [
                summary.get("total_medicines", 0),
                f"{summary.get('total_ordered', 0):,.0f}",
                f"{summary.get('total_sold', 0):,.0f}",
                f"{summary.get('total_remaining', 0):,.0f}",
                f"{summary.get('total_shortage', 0):,.0f}",
                f"{summary.get('total_leftover', 0):,.0f}",
                summary.get("medicines_with_shortage", 0),
                summary.get("medicines_with_leftover", 0),
                f"{summary.get('sold_percentage', 0):.1f}%",
            ],
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        
        # Sheet 3: Issues (if any)
        if len(issues_df) > 0:
            # Flatten row_ref and raw_snippet for Excel
            issues_export = issues_df.copy()
            if "row_ref" in issues_export.columns:
                issues_export["row_source"] = issues_export["row_ref"].apply(
                    lambda x: x.get("source", "") if isinstance(x, dict) else ""
                )
                issues_export["row_number"] = issues_export["row_ref"].apply(
                    lambda x: x.get("row_number", "") if isinstance(x, dict) else ""
                )
                issues_export = issues_export.drop(columns=["row_ref"], errors="ignore")
            
            if "raw_snippet" in issues_export.columns:
                issues_export = issues_export.drop(columns=["raw_snippet"], errors="ignore")
            
            issues_export.to_excel(writer, sheet_name="Issues", index=False)
        else:
            pd.DataFrame(columns=["rule_id", "severity", "medicine_key", "details"]).to_excel(
                writer, sheet_name="Issues", index=False
            )
    
    logger.info(f"Created Excel report: {output_path}")

